import os
import openpyxl
import re
import csv


class Process:
    def __init__(self):
        pass

    def wm_portal_products(self, f):
        wb = openpyxl.load_workbook(filename=f)
        wb_rv = openpyxl.Workbook()

        ws = wb.active
        ws_rv = wb_rv.active
        ws_rv.title = ws.title

        for r, row in enumerate(ws.iter_rows(min_row=7, max_row=ws.max_row), 1):
            for c, column in enumerate(row, 1):
                ws_rv.cell(row=r, column=c, value=column.value)

        wb_rv.save(filename='WMProducts-Current.xlsx')

    def wm_sku(self, f):
        wb = openpyxl.load_workbook(filename=f)
        wb_rv = openpyxl.Workbook()

        ws = wb.active
        ws_rv = wb_rv.active
        ws_rv.title = ws.title

        for r, row in enumerate(ws.iter_rows(min_row=7, max_row=ws.max_row), 1):
            for c, column in enumerate(row, 1):
                ws_rv.cell(row=r, column=c, value=column.value)

        wb_rv.save(filename='WMProducts-Filename.xlsx')

    def wm_users(self, f):
        wb = openpyxl.load_workbook(filename=f)
        wb_rv = openpyxl.Workbook()

        ws = wb.active
        ws_rv = wb_rv.active
        ws_rv.title = ws.title

        for r, row in enumerate(ws.iter_rows(min_row=7, max_row=ws.max_row), 1):
            for c, column in enumerate(row, 1):
                ws_rv.cell(row=r, column=c, value=column.value)

        wb_rv.save(filename='PortalUsers-Current.xlsx')

    def wm_pricing(self, f):
        wb = openpyxl.load_workbook(filename=f)
        wb_rv = openpyxl.Workbook()

        ws = wb.active
        ws_rv = wb_rv.active
        ws_rv.title = ws.title

        for r, row in enumerate(ws.iter_rows(min_row=7, max_row=ws.max_row), 1):
            for c, column in enumerate(row, 1):
                ws_rv.cell(row=r, column=c, value=column.value)

        wb_rv.save(filename='WMSupplier.xlsx')

    def fb_pricing(self, f):
        wb = openpyxl.load_workbook(filename=f)
        wb_rv = openpyxl.Workbook()

        ws = wb.active
        ws_rv = wb_rv.active
        ws_rv.title = ws.title

        for r, row in enumerate(ws.iter_rows(min_row=7, max_row=ws.max_row), 1):
            for c, column in enumerate(row, 1):
                ws_rv.cell(row=r, column=c, value=column.value)

        wb_rv.save(filename='FBProducts-Supplier.xlsx')

    def fb_sku(self, f):
        wb = openpyxl.load_workbook(filename=f)
        wb_rv = openpyxl.Workbook()

        ws = wb.active
        ws_rv = wb_rv.active
        ws_rv.title = ws.title

        for r, row in enumerate(ws.iter_rows(min_row=7, max_row=ws.max_row), 1):
            for c, column in enumerate(row, 1):
                ws_rv.cell(row=r, column=c, value=column.value)

        wb_rv.save(filename='FBProducts-SKUs.xlsx')

    def fb_postage(self, f):
        wb = openpyxl.load_workbook(filename=f)
        wb_rv = openpyxl.Workbook()

        ws = wb.active
        ws_rv = wb_rv.active
        ws_rv.title = ws.title

        for r, row in enumerate(ws.iter_rows(min_row=7, max_row=ws.max_row), 1):
            for c, column in enumerate(row, 1):
                ws_rv.cell(row=r, column=c, value=column.value)

        wb_rv.save(filename='FBPostage.xlsx')

    def fb_products(self, f):
        wb = openpyxl.load_workbook(filename=f)
        wb_rv = openpyxl.Workbook()

        ws = wb.active
        ws_rv = wb_rv.active
        ws_rv.title = ws.title

        for r, row in enumerate(ws.iter_rows(min_row=7, max_row=ws.max_row), 1):
            for c, column in enumerate(row, 1):
                ws_rv.cell(row=r, column=c, value=column.value)

        wb_rv.save(filename='FBProducts-Current.xlsx')

    def mmh_portal_products(self, f):
        wb = openpyxl.load_workbook(filename=f)
        wb_rv = openpyxl.Workbook()

        ws = wb.active
        ws_rv = wb_rv.active
        ws_rv.title = ws.title

        for r, row in enumerate(ws.iter_rows(min_row=7, max_row=ws.max_row), 1):
            for c, column in enumerate(row, 1):
                ws_rv.cell(row=r, column=c, value=column.value)

        wb_rv.save(filename='MMH Current Products.xlsx')

    def mmh_sku(self, f):
        wb = openpyxl.load_workbook(filename=f)
        wb_rv = openpyxl.Workbook()

        ws = wb.active
        ws_rv = wb_rv.active
        ws_rv.title = ws.title

        for r, row in enumerate(ws.iter_rows(min_row=7, max_row=ws.max_row), 1):
            for c, column in enumerate(row, 1):
                ws_rv.cell(row=r, column=c, value=column.value)

        wb_rv.save(filename='MMH SKUS.xlsx')

    def mmh_users(self, f):
        wb = openpyxl.load_workbook(filename=f)
        wb_rv = openpyxl.Workbook()

        ws = wb.active
        ws_rv = wb_rv.active
        ws_rv.title = ws.title

        for r, row in enumerate(ws.iter_rows(min_row=7, max_row=ws.max_row), 1):
            for c, column in enumerate(row, 1):
                ws_rv.cell(row=r, column=c, value=column.value)

        wb_rv.save(filename='MMH Current Users.xlsx')

    def mmh_postage(self, f):
        wb = openpyxl.load_workbook(filename=f)
        wb_rv = openpyxl.Workbook()

        ws = wb.active
        ws_rv = wb_rv.active
        ws_rv.title = ws.title

        for r, row in enumerate(ws.iter_rows(min_row=7, max_row=ws.max_row), 1):
            for c, column in enumerate(row, 1):
                ws_rv.cell(row=r, column=c, value=column.value)

        wb_rv.save(filename='MMH Postage.xlsx')

    def mmh_ticket(self, f):
        wb = openpyxl.load_workbook(filename=f)
        wb_rv = openpyxl.Workbook()

        ws = wb.active
        ws_rv = wb_rv.active
        ws_rv.title = ws.title

        for r, row in enumerate(ws.iter_rows(min_row=7, max_row=ws.max_row), 1):
            for c, column in enumerate(row, 1):
                ws_rv.cell(row=r, column=c, value=column.value)

        wb_rv.save(filename='MMHJobTicketInstructions.xlsx')

    def mmh_group(self, f):
        wb_rv = openpyxl.Workbook()

        ws_rv = wb_rv.active
        ws_rv.title = 'UserGroup'

        with open(f, 'r', newline='\r\n', encoding='utf-16-be') as ug:
            csvr = csv.DictReader(ug, ['Group ID', 'Name', 'Description'])
            for n, line in enumerate(csvr, 1):
                # print(n, line['Group ID'], line['Name'])
                ws_rv.cell(row=n, column=1, value=line['Group ID'])
                ws_rv.cell(row=n, column=2, value=line['Name'])
                ws_rv.cell(row=n, column=3, value=line['Description'])

        wb_rv.save(filename='MMH User Groups.xlsx')


def process_mmh(process_path):
    products = re.compile("(PortalProducts)[\s\S]*.xlsx")
    users = re.compile("(PortalUsers)[\s\S]*.xlsx")
    postage = re.compile("(Postage)[\s\S]*.xlsx")
    sku = re.compile("(SKUs)[\s\S]*.xlsx")
    ticket = re.compile("(JobTicketInstructions)[\s\S]*.xlsx")
    group = re.compile("(UserGroup)[\s\S]*.csv")

    process = Process()

    mmh_files = os.listdir(process_path)

    for f in mmh_files:
        if postage.match(f):
            print("MMH Postage")
            process.mmh_postage(os.path.join(process_path, f))

        if users.match(f):
            print("MMH Portal Users")
            process.mmh_users(os.path.join(process_path, f))

        if sku.match(f):
            print("MMH SKUs")
            process.mmh_sku(os.path.join(process_path, f))

        if products.match(f):
            print("MMH Portal Products")
            process.mmh_portal_products(os.path.join(process_path, f))

        if ticket.match(f):
            print("MMH Job Ticket Instructions")
            process.mmh_ticket(os.path.join(process_path, f))

        if group.match(f):
            print("MMH User Groups")
            process.mmh_group(os.path.join(process_path, f))


def process_wm(process_path):
    products = re.compile("(PortalProducts)[\s\S]*.xlsx")
    users = re.compile("(PortalUsers)[\s\S]*.xlsx")
    pricing = re.compile("(Pricing)[\s\S]*.xlsx")
    sku = re.compile("(SKUs)[\s\S]*.xlsx")

    process = Process()

    wm_files = os.listdir(process_path)
    for f in wm_files:
        if pricing.match(f):
            print("WM Pricing")
            process.wm_pricing(os.path.join(process_path, f))

        if users.match(f):
            print("WM Portal Users")
            process.wm_users(os.path.join(process_path, f))

        if sku.match(f):
            print("WM SKUs")
            process.wm_sku(os.path.join(process_path, f))

        if products.match(f):
            print("WM Portal Products")
            process.wm_portal_products(os.path.join(process_path, f))


def process_fb(process_path):
    products = re.compile("(PortalProducts)[\s\S]*.xlsx")
    postage = re.compile("(Postage)[\s\S]*.xlsx")
    pricing = re.compile("(Pricing)[\s\S]*.xlsx")
    sku = re.compile("(SKUs)[\s\S]*.xlsx")

    process = Process()

    fb_files = os.listdir(process_path)
    for f in fb_files:
        if pricing.match(f):
            print("FB Pricing")
            process.fb_pricing(os.path.join(process_path, f))

        if postage.match(f):
            print("FB Postage")
            process.fb_postage(os.path.join(process_path, f))

        if sku.match(f):
            print("FB SKUs")
            process.fb_sku(os.path.join(process_path, f))

        if products.match(f):
            print("FB Portal Products")
            process.fb_products(os.path.join(process_path, f))


def main():
    # wm_path = os.path.join(os.curdir, 'WM')
    # fb_path = os.path.join(os.curdir, 'FB')
    # mmh_path = os.path.join(os.curdir, 'MMH')
    version = int(input("Daily Web Order Data Process "
                        "(0 for Wellmark, 1 for Farm Bureau, 2 for Medica): "))

    if version == 0:
        process_wm(os.curdir)
    if version == 1:
        process_fb(os.curdir)
    if version == 2:
        process_mmh(os.curdir)


if __name__ == '__main__':
    main()
